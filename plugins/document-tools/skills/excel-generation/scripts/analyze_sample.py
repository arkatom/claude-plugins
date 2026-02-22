#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
サンプルExcel解析スクリプト

サンプルExcelファイルから体裁情報（列幅、フォント、色、配置、罫線等）を
完全に抽出する。スタイル判定は行わず、純粋に体裁情報のみを取得。
"""

import openpyxl
from openpyxl.utils import get_column_letter


def analyze_sample_excel(sample_path):
    """
    サンプルExcelを解析し、体裁情報を返す

    Args:
        sample_path: サンプルExcelファイルのパス

    Returns:
        dict: 体裁情報（シート構造、列幅、セルスタイル等）
    """
    wb = openpyxl.load_workbook(sample_path)

    result = {
        'sheets': []
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        sheet_info = {
            'name': sheet_name,
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'column_widths': {},
            'row_heights': {},
            'freeze_panes': None,
            'cells': {}
        }

        # 列幅を取得
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            width = ws.column_dimensions[col_letter].width
            if width is not None:
                sheet_info['column_widths'][col_idx] = width

        # 行の高さを取得
        for row_idx in range(1, ws.max_row + 1):
            height = ws.row_dimensions[row_idx].height
            if height is not None:
                sheet_info['row_heights'][row_idx] = height

        # フリーズペインを取得
        if ws.freeze_panes:
            sheet_info['freeze_panes'] = ws.freeze_panes

        # セル情報を取得（最初の100行のみ、効率のため）
        for row_idx in range(1, min(101, ws.max_row + 1)):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # セルに値があるか、体裁が設定されている場合のみ記録
                if cell.value or cell.font.bold or cell.fill.start_color.rgb:
                    cell_info = {
                        'value': str(cell.value) if cell.value else '',
                        'font': {
                            'name': cell.font.name,
                            'size': cell.font.size,
                            'bold': cell.font.bold,
                            'color': None
                        },
                        'fill': {
                            'rgb': None
                        },
                        'alignment': {
                            'horizontal': cell.alignment.horizontal,
                            'vertical': cell.alignment.vertical,
                            'wrap_text': cell.alignment.wrap_text
                        },
                        'border': {
                            'left': cell.border.left.style if cell.border.left else None,
                            'right': cell.border.right.style if cell.border.right else None,
                            'top': cell.border.top.style if cell.border.top else None,
                            'bottom': cell.border.bottom.style if cell.border.bottom else None
                        }
                    }

                    # フォント色を取得
                    if cell.font.color:
                        if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                            cell_info['font']['color'] = cell.font.color.rgb
                        elif hasattr(cell.font.color, 'theme'):
                            cell_info['font']['color'] = f"theme{cell.font.color.theme}"

                    # 背景色を取得
                    if cell.fill.start_color:
                        if hasattr(cell.fill.start_color, 'rgb') and cell.fill.start_color.rgb:
                            cell_info['fill']['rgb'] = cell.fill.start_color.rgb

                    sheet_info['cells'][(row_idx, col_idx)] = cell_info

        result['sheets'].append(sheet_info)

    return result


def print_sample_analysis(sample_path):
    """
    サンプルExcelの解析結果を表示（デバッグ用）

    Args:
        sample_path: サンプルExcelファイルのパス
    """
    result = analyze_sample_excel(sample_path)

    print(f"=== サンプルExcel解析結果 ===")
    print()

    for sheet in result['sheets']:
        print(f"シート: {sheet['name']}")
        print(f"  最大行: {sheet['max_row']}, 最大列: {sheet['max_column']}")
        print(f"  列幅: {len(sheet['column_widths'])}列分")
        print(f"  行の高さ: {len(sheet['row_heights'])}行分")
        print(f"  フリーズペイン: {sheet['freeze_panes']}")
        print(f"  セル情報: {len(sheet['cells'])}セル分")
        print()


if __name__ == '__main__':
    # テスト実行
    import sys
    import os

    # プロジェクトルートからの相対パスで指定
    # このスクリプトは .claude/skills/excel-generation/ に配置
    sample_path = os.path.join(
        os.path.dirname(__file__),
        '../../..',  # プロジェクトルートへ
        'docs/specs/_original_specs/Cron処理サンプル_1行のみ残した形.xlsx'
    )

    print(f"サンプルファイル: {sample_path}")
    print()

    if os.path.exists(sample_path):
        print_sample_analysis(sample_path)
    else:
        print(f"エラー: サンプルファイルが見つかりません: {sample_path}")
