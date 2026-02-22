#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel生成エンジン

サンプルExcelの体裁を完全に再現し、スタイル変換を適用して
新しいExcelファイルを生成する汎用ライブラリ。
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter


def apply_cell_style(cell, style_dict):
    """
    セルに体裁を適用

    Args:
        cell: openpyxlのセルオブジェクト
        style_dict: 体裁情報の辞書
    """
    if not style_dict:
        return

    # フォントを適用
    if 'font' in style_dict:
        font_info = style_dict['font']
        color = None

        if font_info.get('color'):
            color_val = font_info['color']
            if isinstance(color_val, str) and color_val.startswith('FF'):
                # RGB形式
                color = color_val
            elif isinstance(color_val, str) and color_val.startswith('theme'):
                # Theme形式は白文字（FFFFFF）として扱う
                color = 'FFFFFF'

        cell.font = Font(
            name=font_info.get('name', 'Meiryo'),
            size=font_info.get('size', 11),
            bold=font_info.get('bold', False),
            color=color
        )

    # 背景色を適用
    if 'fill' in style_dict and style_dict['fill'].get('rgb'):
        rgb = style_dict['fill']['rgb']
        if isinstance(rgb, str):
            # RGB形式の文字列をColorオブジェクトに変換
            cell.fill = PatternFill(
                start_color=Color(rgb=rgb),
                end_color=Color(rgb=rgb),
                fill_type='solid'
            )

    # 配置を適用
    if 'alignment' in style_dict:
        align_info = style_dict['alignment']
        cell.alignment = Alignment(
            horizontal=align_info.get('horizontal'),
            vertical=align_info.get('vertical'),
            wrap_text=align_info.get('wrap_text', False)
        )

    # 罫線を適用
    if 'border' in style_dict:
        border_info = style_dict['border']
        cell.border = Border(
            left=Side(style=border_info.get('left') or 'thin'),
            right=Side(style=border_info.get('right') or 'thin'),
            top=Side(style=border_info.get('top') or 'thin'),
            bottom=Side(style=border_info.get('bottom') or 'thin')
        )


def apply_sheet_format(ws, sheet_template):
    """
    シートに体裁を適用

    Args:
        ws: openpyxlのワークシートオブジェクト
        sheet_template: analyze_sample_excel()から得たシート情報
    """
    # 列幅を適用
    for col_idx, width in sheet_template.get('column_widths', {}).items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # 行の高さを適用
    for row_idx, height in sheet_template.get('row_heights', {}).items():
        ws.row_dimensions[row_idx].height = height

    # フリーズペインを適用
    if sheet_template.get('freeze_panes'):
        ws.freeze_panes = sheet_template['freeze_panes']


def create_excel_with_template(data_rows, template_info, sheet_name='Sheet1'):
    """
    テンプレート情報を使ってExcelを生成

    Args:
        data_rows: データ行のリスト
                   例: [
                       {'SEQ': 1, '処理名': 'メール配信', ...},
                       {'SEQ': 2, '処理名': 'バウンス受信', ...}
                   ]
        template_info: analyze_sample_excel()の返り値
        sheet_name: シート名

    Returns:
        openpyxl.Workbook: 生成されたワークブック
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # テンプレートの最初のシートを使用
    if not template_info['sheets']:
        raise ValueError("テンプレート情報にシートが含まれていません")

    sheet_template = template_info['sheets'][0]

    # シート全体の体裁を適用
    apply_sheet_format(ws, sheet_template)

    # テンプレートのセル情報からタイトル行・ヘッダー行を特定
    # Row 2: タイトル
    # Row 4: ヘッダー
    # Row 5以降: データ

    # テンプレートのセルスタイルを適用
    for (row_idx, col_idx), cell_style in sheet_template.get('cells', {}).items():
        if row_idx <= 4:  # タイトル・空行・ヘッダー行
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = cell_style.get('value')
            apply_cell_style(cell, cell_style)

    # データ行を追加（Row 5から開始）
    if data_rows:
        headers = list(data_rows[0].keys()) if data_rows else []

        # ヘッダー行のスタイルを取得（Row 4）
        header_style = sheet_template.get('cells', {}).get((4, 1), {})

        for data_idx, data_row in enumerate(data_rows):
            row_idx = 5 + data_idx
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = data_row.get(header, '')

                # データ行のスタイルを適用（Row 5のスタイルを参照）
                data_style = sheet_template.get('cells', {}).get((5, col_idx), {})
                if data_style:
                    apply_cell_style(cell, data_style)

    return wb


def create_excel_default(data_rows, sheet_name='Sheet1', style='B'):
    """
    デフォルト体裁でExcelを生成

    Args:
        data_rows: データ行のリスト
        sheet_name: シート名
        style: スタイル ('A', 'B', 'H')

    Returns:
        openpyxl.Workbook: 生成されたワークブック
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # デフォルト体裁の定義
    title_format = {
        'font': Font(name='Meiryo', size=20, bold=True),
        'alignment': Alignment(horizontal='left', vertical='center')
    }

    header_format = {
        'font': Font(name='Meiryo', size=12, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='C22114', end_color='C22114', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

    cell_format = {
        'font': Font(name='Meiryo', size=11),
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

    # Row 2: タイトル
    ws.cell(row=2, column=1, value=f'■{sheet_name}')
    ws.cell(row=2, column=1).font = title_format['font']
    ws.cell(row=2, column=1).alignment = title_format['alignment']
    ws.row_dimensions[2].height = 25

    # Row 4: ヘッダー
    if data_rows:
        headers = list(data_rows[0].keys())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_format['font']
            cell.fill = header_format['fill']
            cell.alignment = header_format['alignment']
            cell.border = header_format['border']

        # Row 5以降: データ
        for data_idx, data_row in enumerate(data_rows):
            row_idx = 5 + data_idx
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=data_row.get(header, ''))
                cell.font = cell_format['font']
                cell.alignment = cell_format['alignment']
                cell.border = cell_format['border']

    # 列幅の自動調整（デフォルト値）
    default_widths = {
        1: 6.33,   # SEQ列
        2: 12.5,   # 画面区分列
        3: 66.5,   # Cron設定列
        4: 18.66,  # 実行タイミング列
        5: 113.33  # 処理内容列
    }

    for col_idx, width in default_widths.items():
        if col_idx <= (len(data_rows[0].keys()) if data_rows else 0):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

    # フリーズペイン
    ws.freeze_panes = 'A5'

    return wb


if __name__ == '__main__':
    # テスト実行
    test_data = [
        {
            'SEQ': 1,
            '画面区分': 'メール配信',
            'Cron設定': '*/15 * * * *',
            '実行タイミング': '15分おき',
            '処理内容': 'メール一斉配信処理'
        },
        {
            'SEQ': 2,
            '画面区分': 'バウンス受信',
            'Cron設定': '*/5 * * * *',
            '実行タイミング': '5分おき',
            '処理内容': 'バウンスメール受信処理'
        }
    ]

    wb = create_excel_default(test_data, sheet_name='Cron処理一覧', style='B')
    wb.save('test_output.xlsx')
    print('テストExcelを生成しました: test_output.xlsx')
