#!/usr/bin/env python3
"""
review.json から Excel 形式のテストケースを生成するスクリプト
1ステップ = 1行の実務的なフォーマット

使用方法:
  cat review.json | python3 generate_excel.py output.xlsx

出力形式:
  - 1テストケース × Nステップ = N行
  - 画面・機能ごとにシートを分割
  - 優先度で行を色分け（Critical=赤、High=オレンジ、Medium=黄、Low=緑）
  - ヘッダー行を固定
  - フィルター機能を有効化
  - 列幅を自動調整
  - 結果列にドロップダウンリスト（OK/NG/PN/NA）を設定

列構成:
  テストケースID | ステップ番号 | 優先度 | リスクスコア | カテゴリ | サブカテゴリ |
  タイトル | 前提 | 操作 | 期待挙動 | 結果 | 実施日 | 実施者 | 備考

依存関係:
  pip install openpyxl
"""

import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def get_priority_color(priority):
    """優先度に応じた背景色を返す"""
    colors = {
        'Critical': 'FFCCCC',  # 薄い赤
        'High': 'FFCC99',      # 薄いオレンジ
        'Medium': 'FFFFCC',    # 薄い黄
        'Low': 'CCFFCC'        # 薄い緑
    }
    return colors.get(priority, 'FFFFFF')


def auto_adjust_column_width(worksheet):
    """列幅を自動調整"""
    column_widths = {
        'A': 15,  # テストケースID
        'B': 8,   # ステップ番号
        'C': 10,  # 優先度
        'D': 12,  # リスクスコア
        'E': 12,  # カテゴリ
        'F': 18,  # サブカテゴリ
        'G': 35,  # タイトル
        'H': 45,  # 前提
        'I': 55,  # 操作
        'J': 55,  # 期待挙動
        'K': 8,   # 結果
        'L': 12,  # 実施日
        'M': 12,  # 実施者
        'N': 25   # 備考
    }

    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width


def extract_preconditions(testcase):
    """
    前提条件を抽出（具体的な条件のみ）

    原則: 「システムが正常に稼働している」のような一般的すぎる条件は記載しない。
    テスト固有の前提条件のみを記載する。
    """
    preconditions = []

    # 明示的な前提条件があればそれを使用
    if 'preconditions' in testcase and testcase['preconditions']:
        if isinstance(testcase['preconditions'], list):
            return '\n'.join(testcase['preconditions'])
        else:
            return testcase['preconditions']

    # テストケースの内容から具体的な前提条件を推定
    title = testcase.get('title', '')
    description = testcase.get('description', '')
    category = testcase.get('category', '')
    subcategory = testcase.get('subcategory', '')

    # セミナー申込・お問い合わせフォーム関連
    if '送信' in description or '完了' in description or 'メール' in title:
        preconditions.append('説明会日程CSVファイルに有効な日程が登録されている')

    # reCAPTCHA関連
    if 'recaptcha' in title.lower() or 'captcha' in description.lower():
        preconditions.append('reCAPTCHAが有効である（NOCAPTCHA_SECRET設定済み）')

    # 認証関連
    if 'ログイン' in title or '認証' in description:
        preconditions.append('ユーザーがログアウト状態である')
    elif 'ログアウト' in title:
        preconditions.append('ユーザーがログイン済みである')

    # 管理画面関連
    if 'admin' in title.lower() or '管理画面' in description:
        preconditions.append('管理者としてログイン済みである')

    # CSVファイル関連
    if 'CSV' in title or 'csv' in description:
        if '空' in title or '存在しない' in title:
            pass  # 意図的に空または削除するテストなので前提不要
        else:
            preconditions.append('説明会日程CSVファイルが存在する')

    # 日程フィルタリング関連
    if '4日前' in description or '日程' in title and 'フィルタ' in subcategory:
        preconditions.append('説明会日程CSVファイルに過去・現在・未来の日程が含まれている')

    # 前提条件がない場合は空文字（一般的な条件は書かない）
    return '\n'.join(preconditions) if preconditions else ''


def create_testcase_sheet(workbook, sheet_name, testcases, target_name):
    """テストケースシートを作成（1ステップ=1行）"""
    ws = workbook.create_sheet(title=sheet_name)

    # ヘッダー行
    headers = [
        'テストケースID', 'ステップ番号', '優先度', 'リスクスコア',
        'カテゴリ', 'サブカテゴリ', 'タイトル', '前提', '操作', '期待挙動',
        '結果', '実施日', '実施者', '備考'
    ]

    # ヘッダー行を書き込み
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 罫線
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        cell.border = thin_border

    # 結果列のドロップダウンリストを設定
    dv = DataValidation(type="list", formula1='"OK,NG,PN,NA"', allow_blank=True)
    dv.error = '選択肢から選んでください: OK, NG, PN, NA'
    dv.errorTitle = '入力エラー'
    dv.prompt = 'OK: 合格, NG: 不合格, PN: Pending, NA: Not Applicable'
    dv.promptTitle = '結果を選択'
    ws.add_data_validation(dv)

    # データ行（1ステップ=1行）
    current_row = 2

    for tc in testcases:
        tc_id = tc.get('id', '')
        priority = tc.get('priority', '')
        risk_score = tc.get('risk_score', 0)
        category = tc.get('category', '')
        subcategory = tc.get('subcategory', '')
        title = tc.get('title', '')

        # 前提条件（最初のステップのみ）
        preconditions = extract_preconditions(tc)

        # ステップと期待結果を取得
        steps = tc.get('steps', [])
        expected_results = tc.get('expected_results', [])

        # ステップがない場合は1行だけ出力（サマリー行）
        if not steps:
            steps = ['（テストステップ未定義）']
            expected_results = ['（期待結果未定義）']

        # 各ステップを1行ずつ出力
        for step_num, (step, expected) in enumerate(zip(steps, expected_results), 1):
            row_data = [
                tc_id,                                    # テストケースID
                step_num,                                 # ステップ番号
                priority,                                 # 優先度
                risk_score,                               # リスクスコア
                category,                                 # カテゴリ
                subcategory,                              # サブカテゴリ
                title,                                    # タイトル
                preconditions if step_num == 1 else '-',  # 前提（最初のステップのみ）
                step,                                     # 操作
                expected,                                 # 期待挙動
                '',                                       # 結果（空欄）
                '',                                       # 実施日
                '',                                       # 実施者
                ''                                        # 備考
            ]

            # セルに書き込み
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num, value=value)
                cell.alignment = Alignment(vertical='top', wrap_text=True)

                # 罫線
                thin_border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
                cell.border = thin_border

            # 結果列にドロップダウンリストを追加
            result_cell = ws.cell(row=current_row, column=11)  # K列（結果）
            dv.add(result_cell)

            # 優先度に応じて行を色分け（全列）
            fill_color = get_priority_color(priority)
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=current_row, column=col_num).fill = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type='solid'
                )

            current_row += 1

    # ヘッダー行を固定
    ws.freeze_panes = 'A2'

    # フィルター機能を有効化
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{current_row - 1}'

    # 列幅を自動調整
    auto_adjust_column_width(ws)


def create_summary_sheet(workbook, data):
    """サマリーシートを作成"""
    ws = workbook.active
    ws.title = '概要'

    # タイトル
    ws['A1'] = 'テストケース生成サマリー'
    ws['A1'].font = Font(size=16, bold=True, color='1F4E78')

    # メタデータ
    meta = data.get('meta', {})
    target = meta.get('target', '')
    generated_at = meta.get('generated_at', '')
    total = meta.get('total_testcases', 0)

    row = 3
    ws[f'A{row}'] = '対象'
    ws[f'B{row}'] = target
    ws[f'A{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = '生成日時'
    ws[f'B{row}'] = generated_at
    ws[f'A{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = '総テストケース数'
    ws[f'B{row}'] = total
    ws[f'A{row}'].font = Font(bold=True)

    # 優先度別統計
    row += 2
    ws[f'A{row}'] = '優先度別内訳'
    ws[f'A{row}'].font = Font(size=14, bold=True)

    by_priority = meta.get('by_priority', {})
    for priority in ['Critical', 'High', 'Medium', 'Low']:
        row += 1
        count = by_priority.get(priority, 0)
        ws[f'A{row}'] = priority
        ws[f'B{row}'] = f'{count}件'
        ws[f'A{row}'].font = Font(bold=True)

        # 優先度の色を適用
        fill_color = get_priority_color(priority)
        ws[f'A{row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

    # 結果コードの説明
    row += 3
    ws[f'A{row}'] = '結果コードの説明'
    ws[f'A{row}'].font = Font(size=14, bold=True)

    result_codes = [
        ('OK', '合格', '66FF66'),
        ('NG', '不合格', 'FF6666'),
        ('PN', 'Pending（未実施・保留）', 'FFFF99'),
        ('NA', 'Not Applicable（該当なし・スキップ）', 'CCCCCC')
    ]

    for code, desc, color in result_codes:
        row += 1
        ws[f'A{row}'] = code
        ws[f'B{row}'] = desc
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    # 列幅を調整
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40


def main():
    """メイン処理"""
    try:
        # コマンドライン引数から出力ファイル名を取得
        if len(sys.argv) < 2:
            print("使用方法: cat review.json | python3 generate_excel.py output.xlsx", file=sys.stderr)
            sys.exit(1)

        output_file = sys.argv[1]

        # stdin から review.json を読み込み
        data = json.load(sys.stdin)

        # testcases キーまたは test_perspectives キーを探す（互換性のため）
        testcases_key = 'testcases' if 'testcases' in data else 'test_perspectives'
        all_testcases = data.get(testcases_key, [])

        # Excelワークブックを作成
        wb = Workbook()

        # サマリーシートを作成
        create_summary_sheet(wb, data)

        # meta.target を使用して、画面名を取得
        target_name = data.get('meta', {}).get('target', 'テストケース')

        # シート名を最大31文字に制限（Excel制約）
        sheet_name = target_name[:31] if len(target_name) > 31 else target_name

        # すべてのテストケースを1つのシートに出力
        create_testcase_sheet(wb, sheet_name, all_testcases, target_name)

        # Excelファイルを保存
        wb.save(output_file)

        # 統計情報を出力
        total_steps = sum(len(tc.get('steps', [])) for tc in all_testcases)

        print(f"✅ Excelファイルを生成しました: {output_file}", file=sys.stderr)
        print(f"   - 総テストケース数: {len(all_testcases)}件", file=sys.stderr)
        print(f"   - 総ステップ数: {total_steps}行", file=sys.stderr)
        print(f"   - シート数: 2（概要 + {sheet_name}）", file=sys.stderr)
        print(f"   - フォーマット: 1ステップ=1行（実務標準）", file=sys.stderr)
        print(f"   - 結果列: ドロップダウンリスト（OK/NG/PN/NA）", file=sys.stderr)

    except ImportError:
        print("エラー: openpyxl がインストールされていません。", file=sys.stderr)
        print("インストール方法: pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"エラー: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
